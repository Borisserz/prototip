import io
import json
import logging

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger("ExcelRenderer")


class ExcelRenderer:
    @staticmethod
    def render_json_to_excel(data_json: str, title: str = "Аналитический Отчет") -> bytes:
        """
        Преобразует JSON (список словарей) в красиво оформленный Excel файл.
        Возвращает байты (binary content).
        """
        try:
            data = json.loads(data_json)
            if not isinstance(data, list) or len(data) == 0:
                raise ValueError("Data must be a non-empty JSON list of objects.")

            df = pd.DataFrame(data)

            wb = Workbook()
            ws = wb.active
            ws.title = "Данные"

            # Добавляем заголовок
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(size=14, bold=True, color="FFFFFF")
            title_cell.fill = PatternFill(
                start_color="4F46E5", end_color="4F46E5", fill_type="solid"
            )  # Indigo-600
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            # Стили для заголовков столбцов
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="1E293B", end_color="1E293B", fill_type="solid"
            )  # Slate-800
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Заполняем заголовки
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(
                    row=2, column=col_num, value=column_title.replace("_", " ").capitalize()
                )
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Данные
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 3):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.number_format = "#,##0.00"

            # Автоширина столбцов
            for col in ws.columns:
                # Skip merged cells for determining column width, find the first non-merged cell
                valid_cell = None
                for cell in col:
                    if hasattr(cell, "column_letter"):
                        valid_cell = cell
                        break

                if not valid_cell:
                    continue

                column = valid_cell.column_letter
                max_length = 0
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = min(adjusted_width, 50)  # Max width 50

            # Сохраняем в память
            excel_io = io.BytesIO()
            wb.save(excel_io)
            excel_io.seek(0)
            return excel_io.read()

        except Exception as e:
            logger.error(f"Error rendering Excel: {e}")
            raise
